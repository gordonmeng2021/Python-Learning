import openpyxl
import os
import pandas as pd
os.getcwd()

file = 'Excel extract data/instagram_report.xlsx'
data =  pd.ExcelFile(file)
 #print(data.sheet_names) this print the name of the sheet

df = data.parse('Socialinsider | Analytics Tool')
df.info
df.head(10)


# *****-------read in the spreadsheet data

ps = openpyxl.load_workbook('Excel extract data/instagram_report.xlsx') # relative path
sheet = ps['Socialinsider | Analytics Tool'] # sheet name

# print(sheet.max_row)
print()
print("------Key Performance Indicators------")
print()

for row in range(2,sheet.max_row +1):
    result = sheet['A'+str(row)].value
    if result == 'Key Performance Indicators':
        for i in range(24):
            topic = sheet['A'+str(i+8)].value
            
            current = sheet['B'+str(i+8)].value
            
            previous =sheet['C'+str(i+8)].value
            change =sheet['D'+str(i+8)].value
            print(topic,"---- Current :",current,"Previous :",previous,"Change :",change)
            
            
            
            
        
    