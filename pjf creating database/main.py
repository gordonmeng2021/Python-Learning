
from selenium.webdriver.common.by import By
import selenium
from selenium import webdriver
import time
from selenium.webdriver.support.select import Select
import openpyxl
import os
import pandas as pd
from openpyxl.styles import Font
import openpyxl
from openpyxl import Workbook
import re
from selenium.webdriver.common.desired_capabilities import DesiredCapabilities
from googletrans import Translator
import docx

translator = Translator()

workbook = Workbook()
sheet = workbook.active
sheet["A1"] = "Name"
sheet["A1"].font = Font(bold=True)
sheet["B1"] = "产品名称"
sheet["B1"].font = Font(bold=True)
sheet["C1"] = "Winery"
sheet["C1"].font = Font(bold=True)
sheet["D1"] = "酒庄（生产厂家)"
sheet["D1"].font = Font(bold=True)
sheet["E1"] = "Vintage"
sheet["E1"].font = Font(bold=True)
sheet["F1"] = "产品描述"
sheet["F1"].font = Font(bold=True)
sheet["G1"] = "Category"
sheet["G1"].font = Font(bold=True)
sheet["H1"] = "Country of Origin"
sheet["H1"].font = Font(bold=True)
sheet["I1"] = "出产国"
sheet["I1"].font = Font(bold=True)
sheet["J1"] = "Region 1"
sheet["J1"].font = Font(bold=True)
sheet["K1"] = "葡萄产区1"
sheet["K1"].font = Font(bold=True)
sheet["L1"] = "Region 2"
sheet["L1"].font = Font(bold=True)
sheet["M1"] = "葡萄产区2"
sheet["M1"].font = Font(bold=True)
sheet["N1"] = "Graphes"
sheet["N1"].font = Font(bold=True)
sheet["O1"] = "葡萄品种"
sheet["O1"].font = Font(bold=True)
sheet["P1"] = "Alcohol Content"
sheet["P1"].font = Font(bold=True)

# Region2= ['Aconcagua Valley', 'Aloxe-Corton', 'Alpilles', 'Alsace Grand Cru', 'Amarone della Valpolicella', 'Anjou', 'Arbois', 'Auckland', 'Auxey-Duresses', 'Ballard Canyon', 'Bandol', 'Barbaresco', "Barbera d'Alba", "Barbera d'Asti", 'Barolo', 'Barsac', 'Bas-Armagnac', 'Beaujolais-Villages', 'Beaune', 'Bellet', 'Blagny', 'Boca', 'Bolgheri', 'Bordeaux Superieur', 'Bourgogne Aligoté', 'Bourgogne Passe-tout-grains', 'Brouilly', 'Brunello di Montalcino', 'Cahors', 'Calvados', 'Campo de Borja', 'Central Coast', 'Central Otago', 'Chablis', 'Chambolle-Musigny', "Chambolle-Musigny 1er Cru 'Les Amoureuses'", 'Chassagne-Montrachet', 'Chateauneuf-du-Pape', 'Chianti', 'Chiroubles', 'Château-Chalon', 'Clos de Tart Grand Cru', 'Clos de la Roche Grand Cru', "Colli dell'Etruria Centrale", 'Collines Rhodaniennes', 'Cornas', 'Corton-Charlemagne Grand Cru', 'Cote Chalonnaise', 'Cote Rotie', 'Cote de Beaune', 'Cote de Nuits', 'Coteaux Bourguignons', 'Cremant de Bourgogne', 'Crozes Hermitage', "Crémant d'Alsace", 'Crémant de Bourgogne', 'Cuna', 'Côte de Brouilly', 'Côte de Nuits Villages', "Côte-d'Or", 'Côtes de Bordeaux', 'Côtes du Jura', 'Côtes-de-Provence', 'Côtes-du-Rhône', "Dolcetto d'Alba", 'Dry Creek Valley', 'Entre-Deux-Mers', 'Fixin', 'Fleurie', 'Gattinara', 'Gevrey-Chambertin', 'Gigondas', 'Givry', 'Grand Champagne', 'Grands-Échezeaux Grand Cru', 'Graves', 'Greco di Tufo', 'Griotte-Chambertin Grand Cru', 'Haut-Medoc', 'Hautes Cotes de Beaune', "Hawke's Bay", 'Hermitage', 'Hérault', 'Irpinia', 'Juliénas', 'Ladoix', 'Lalande de Pomerol', 'Langhe', 'Languedoc', 'Limoux', 'Maconnais', 'Maipo Valley', 'Maremma Toscana', 'Margaux', 'Marlborough', 'Marsannay', 'Maury', 'Medoc', 'Mercurey', 'Meursault', 'Montalcino', 'Monthélie', 'Montsant', 'Morey-Saint-Denis', 'Morgon', 'Moulin-à-Vent', 'Moulis', 'Moulis-en-Medoc', 'Mount Etna', 'Napa Valley', "Nebbiolo d'Alba", 'Nizza', 'North Coast', 'Northern Rhône', 'Nuits-Saint-Georges', 'Palette', 'Paso Robles', 'Pauillac', 'Pessac-Léognan', 'Petite Champagne', 'Pomard', 'Pomerol', 'Pouilly-Fuisse', 'Prosecco', 'Puente Alto', 'Puligny-Montrachet', 'Qiu Shan Valley', 'Recioto della Valpolicella Classico', 'Ribera del Duero', 'Richebourg Grand Cru', 'Rio Negro', 'Robertson', 'Roero', 'Rosso di Montalcino', 'Russian River Valley', 'Saint Chinian', 'Saint-Aubin', 'Saint-Emilion', 'Saint-Emilion Grand Cru', 'Saint-Estephe', 'Saint-Joseph', 'Saint-Julien', 'Saint-Veran', "Sant'Antonio", 'Santa Ynez Valley', 'Santenay', 'Saumur-Champigny', 'Sauternes', 'Savigny-les-Beaune', 'Sonoma County', 'Southern Rhône', 'Taurasi', 'Toscana', 'Touraine', 'Valdarno di Sopra', 'Valpolicella', 'Volnay', 'Vosne-Romanee', 'Vougeot', 'Waipara Valley', 'Willamette Valley', 'Échezeaux Grand Cru']
region2_list = []
fail_list = []
testing = []
caps = DesiredCapabilities().CHROME
caps["pageLoadStrategy"] = "eager"
# disable image laoding
chrome_options = webdriver.ChromeOptions()
prefs = {"profile.managed_default_content_settings.images": 2}
chrome_options.add_experimental_option("prefs", prefs)
driver = webdriver.Chrome(
    desired_capabilities=caps,
    executable_path="./chromedriver",
    options=chrome_options,
)
name = ""

url_list = [
    # [
    #     "https://www.vivino.com/HK/en/opus-one-opus-one/w/1911534",
    #     "https://wine-world.com/winery/opus-one-winery/20d9c590-2ca8-40ea-a52d-1124f6f00abd",
    # ],
    # [
    #     "https://www.vivino.com/HK/en/sena-aconcagua-valley/w/25474?year=2015",
    #     "https://wine-world.com/winery/sena/cf62c415-0363-45f3-9145-80cd8ceb8952",
    # ],
    # [
    #     "https://www.vivino.com/HK/en/perrier-jouet-grand-brut-champagne/w/79160",
    #     "https://wine-world.com/winery/chateau-duhart-milon/de54f227-2dc9-4b4a-9ddc-868199053be0",
    # ],
    # [
    #     "https://www.vivino.com/HK/en/bollinger-rose-brut-champagne/w/1152435",
    #     "https://wine-world.com/winery/domaine-leflaive/84092B2F-D357-470B-BF1E-878711C000E0",
    # ],
    # [
    #     "https://www.vivino.com/HK/en/domaine-tortochot-chambertin-grand-cru/w/1733384?year=2018",
    #     "https://wine-world.com/winery/domaine-tortochot/3AA44ED3-6E8E-4241-93DC-03A2B46F7144",
    # ],
    # [
    #     "https://www.vivino.com/HK/en/joseph-drouhin-corton-charlemagne-grand-cru/w/1769969",
    #     "https://wine-world.com/winery/maison-joseph-drouhin/7ef8a3cc-fda1-474c-9a0d-791fcc3ce39d",
    # ],
]

doc = docx.Document("pjf.docx")
all_paras = doc.paragraphs

def add_url_list():
    global url_list
    time = 0
    pjf_product_one_wine_list = []
    for para in all_paras:
        pjf_product_one_wine_list.append(remove_comma(para.text))
        time += 1
        if time == 2:
            url_list.append(pjf_product_one_wine_list)
            pjf_product_one_wine_list = []
            time = 0
        

def delete_empty_url_list():
    global url_list
    while (['','']) in url_list:
        url_list.remove(['',''])

# sheet name
def append_region2_list():
    global region2_list
    ok = openpyxl.load_workbook(
        "./Region_07262022.xlsx"
    )  # relative path
    sheet = ok["PM (2)"]
    for i in range(2, sheet.max_row + 1):
        # print(sheet["A"+str(i)].value)
        region = sheet["A" + str(i)].value
        if region != None:
            region2_list.append(sheet["A" + str(i)].value)
        else:
            continue


def find_region_nearby(region):
    ps = openpyxl.load_workbook(
        "./Region_07262022.xlsx"
    )  # relative path
    sheet = ps["PM (2)"]
    big_list = []
    status = ""
    for i in range(2, sheet.max_row + 1):
        value = ""
        temp_list = []
        value = sheet["A" + str(i)].value

        if value == region:
            temp_list.append(value)
            temp_list.append(sheet["B" + str(i)].value)
            temp_list.append(sheet["C" + str(i)].value)
            temp_list.append(sheet["D" + str(i)].value)
            temp_list.append(sheet["E" + str(i)].value)
            temp_list.append(sheet["F" + str(i)].value)
            big_list.append(temp_list)
            break

        else:
            for k in range(2, sheet.max_row + 1):
                value = sheet["C" + str(k)].value
                if value == region:
                    temp_list.append("")
                    temp_list.append("")
                    temp_list.append(value)
                    temp_list.append(sheet["D" + str(k)].value)
                    temp_list.append(sheet["E" + str(k)].value)
                    temp_list.append(sheet["F" + str(k)].value)
                    big_list.append(temp_list)
                    status = "q"
                    break

        if status == "q":
            break
    return big_list


def remove_comma(word):
    new = re.sub(",", "", word)
    return new


def remove_clickno(name):
    y = re.sub("\n.+", "", name)
    return y


def remove_describe_heading(heading):
    y = re.sub(".+\n", "", heading)
    return y


def remove_pencent(graphe):
    x = ""
    x = re.sub("[\d]+% ", "", graphe)
    x = re.sub("\S+/", "", x)
    return x


def check_category(category):
    if category == "Sparkling wine":
        return "Sparkling"
    else:
        return category


# initialize the program

append_region2_list()
add_url_list()
delete_empty_url_list()
print(url_list)

for i in range(len(url_list)):
    vintage = ""
    for k in range(2):
        list = []
        print("Pistion 1")
        driver.get(url_list[i][k])
        print("Pistion 2")

        def find_vintage(name):
            global vintage
            try:
                new = re.search(".[\d]{2,4}$", name)
                vintage = new.group()
            except:
                vintage = ""
                print("It has no vintage")

        def find_category():
            # category=driver.find_element_by_xpath("//*[@id='wine-location-header']/div/div/div/div[1]/div").text
            category = driver.find_element(
                By.XPATH,
                "//*[@id='wine-location-header']/div/div/div/div[1]/div/span[4]/a",
            ).text
            # category = driver.find_element_by_xpath(
            #     "//*[@id='wine-location-header']/div/div/div/div[1]/div/span[4]/a"
            # ).text
            list.append(check_category(category))

        def find_wine_name2():
            global name
            wine_sur_name = driver.find_element(
                By.XPATH,
                "/html/body/div[2]/div[5]/div/div/div[2]/div[1]/div/h1/span[1]/a",
            ).text
            wine_name = driver.find_element(
                By.XPATH,
                "/html/body/div[2]/div[5]/div/div/div[2]/div[1]/div/h1/span[2]",
            ).text

            list.append(wine_sur_name + " " + wine_name) 
            name = wine_sur_name + " " + wine_name
            find_vintage(name)

        #  function need faster
        def scroll_down2():
            driver.execute_script("window.scrollTo(0, 1000);")

        def find_wine_winery2():

            try:
                find_wine_winery = driver.find_element(
                    By.XPATH,
                    "//*[@id='wine-page-lower-section']/div/div[1]/div/div[2]/table/tbody/tr[1]/td/a",
                ).text
                list.append(find_wine_winery)
            except:
                find_wine_winery = driver.find_element(
                    By.XPATH,
                    "//*[@id='wine-page-lower-section']/div[3]/div/div[5]/table/tbod y/tr[1]/td/a",
                ).text

                list.append(find_wine_winery)

        def find_graphes2():
            
            graphes = driver.find_element(
                By.XPATH,
                "//*[@id='wine-page-lower-section']/div/div[1]/div/div[2]/table/tbody/tr[2]/td",
            ).text

            list.append(remove_pencent(graphes))
            chi_name = translator.translate(remove_pencent(graphes), dest="zh-cn").text
            list.append(chi_name)

        def find_region2():

            name = driver.find_element(
                By.XPATH, "//table[contains(@class,'wineFacts__wineFacts')]"
            ).text
            name = re.sub("\n", "", name)
            name = re.sub(".+Region", "", name)
            name = re.sub("Wine.+", "", name)
            name = name.split(" / ")
            print(name[len(name) - 1])
            if name[len(name) - 1] in region2_list:
                list.append(find_region_nearby(name[len(name) - 1]))
            elif name[len(name) - 2] in region2_list:
                list.append(find_region_nearby(name[len(name) - 2]))
            else:
                list.append(find_region_nearby(name[len(name) - 3]))

        def find_alcohol2():
            
            temp = []
            try:
                alcohol = driver.find_element(
                    By.XPATH,
                    "//*[@id='wine-page-lower-section']/div/div[1]/div/div[2]/table/tbody/tr[5]/td/span",
                ).text
                temp.append(alcohol)

            except:
                list.append("N/A %")

            for i in temp:
                if i[-1] == "%":
                    list.append(alcohol)

                else:
                    alcohol = driver.find_element(
                        By.XPATH,
                        "//*[@id='wine-page-lower-section']/div/div[1]/div/div[2]/table/tbody/tr[5]/td/span",
                    ).text

                    for i in alcohol:
                        if i == "%":
                            list.append(alcohol)
                        else:
                            list.append("N/A %")
                            break
                break

        # add bottom from here  ***************************************************

        def find_wine_name():
            global name
            list.clear()
            wine_sur_name = driver.find_element(
                By.XPATH,
                "/html/body/div[2]/div[5]/div/div/div[2]/div[1]/div/h1/span[1]/a",
            ).text
            wine_name = driver.find_element(
                By.XPATH,
                "/html/body/div[2]/div[5]/div/div/div[2]/div[1]/div/h1/span[2]",
            ).text

            list.append(wine_sur_name + " " + wine_name)
            name = wine_sur_name + " " + wine_name
            find_vintage(name)

        def scroll_down():
            # **********faster
            driver.execute_script("window.scrollTo(0, document.body.scrollHeight-900);")

        def find_wine_winery():
            
            try:
                find_wine_winery = driver.find_element(
                    By.XPATH,
                    "//*[@id='wine-page-lower-section']/div[3]/div/div[4]/table/tbody/tr[1]/td/a",
                ).text
                list.append(find_wine_winery)
            except:
                try:
                    find_wine_winery = driver.find_element(
                        By.XPATH,
                        "//*[@id='wine-page-lower-section']/div[3]/div/div[5]/table/tbody/tr[1]/td/a",
                    ).text
                    list.append(find_wine_winery)
                except:
                    find_wine_winery = driver.find_element(
                        By.XPATH,
                        "//*[@id='wine-page-lower-section']/div[4]/div/div[3]/table/tbody/tr[1]/td/a",
                    ).text
                    list.append(find_wine_winery)

        def find_graphes():

            try:
                graphes = driver.find_element(
                    By.XPATH,
                    "//*[@id='wine-page-lower-section']/div[3]/div/div[4]/table/tbody/tr[2]/td",
                ).text
                list.append(remove_pencent(graphes))
                chi_name = translator.translate(
                    remove_pencent(graphes), dest="zh-cn"
                ).text
                list.append(chi_name)

            except:
                try:
                    graphes = driver.find_element(
                        By.XPATH,
                        "//*[@id='wine-page-lower-section']/div[3]/div/div[4]/table/tbody/tr[2]/td",
                    ).text
                    list.append(remove_pencent(graphes))
                    chi_name = translator.translate(
                        remove_pencent(graphes), dest="zh-cn"
                    ).text
                    list.append(chi_name)
                except:
                    try:
                        graphes = driver.find_element(
                            By.XPATH,
                            "//*[@id='wine-page-lower-section']/div[3]/div/div[5]/table/tbody/tr[2]/td",
                        ).text
                        list.append(remove_pencent(graphes))
                        chi_name = translator.translate(
                            remove_pencent(graphes), dest="zh-cn"
                        ).text
                        list.append(chi_name)
                    except:
                        try:
                            graphes = driver.find_element(
                                By.XPATH,
                                "//*[@id='wine-page-lower-section']/div[4]/div/div[3]/table/tbody/tr[2]/td",
                            ).text
                            list.append(remove_pencent(graphes))
                            chi_name = translator.translate(
                                remove_pencent(graphes), dest="zh-cn"
                            ).text
                            list.append(chi_name)
                        except:
                            graphes = driver.find_element(
                                By.XPATH,
                                "//*[@id='wine-page-lower-section']/div[3]/div/div[5]/table/tbody/tr[2]/td/span/a",
                            ).text
                            list.append(remove_pencent(graphes))
                            chi_name = translator.translate(
                                remove_pencent(graphes), dest="zh-cn"
                            ).text
                            list.append(chi_name)

            print(remove_pencent(graphes))

        def find_region():
            name = driver.find_element(
                By.XPATH, "//table[contains(@class,'wineFacts__wineFacts')]"
            ).text
            name = re.sub("\n", "", name)
            name = re.sub(".+Region", "", name)
            name = re.sub("Wine.+", "", name)
            name = name.split(" / ")
            print(name[len(name) - 1])
            if name[len(name) - 1] in region2_list:
                list.append(find_region_nearby(name[len(name) - 1]))
            elif name[len(name) - 2] in region2_list:
                list.append(find_region_nearby(name[len(name) - 2]))
            else:
                list.append(find_region_nearby(name[len(name) - 3]))

        def find_alcohol():
            temp = []
            try:
                alcohol = driver.find_element(
                    By.XPATH,
                    "//*[@id='wine-page-lower-section']/div[3]/div/div[4]/table/tbody/tr[5]/td/span",
                ).text
                temp.append(alcohol)

            except:
                try:
                    alcohol = driver.find_element(
                        By.XPATH,
                        "//*[@id='wine-page-lower-section']/div[3]/div/div[5]/table/tbody/tr[5]/td/span",
                    ).text
                    temp.append(alcohol)
                except:
                    try:

                        alcohol = driver.find_element(
                            By.XPATH,
                            "//*[@id='wine-page-lower-section']/div[4]/div/div[3]/table/tbody/tr[5]/td/span",
                        ).text
                        temp.append(alcohol)
                    except:

                        temp.append("N/A %")

            for i in temp:
                if i[-1] == "%":
                    list.append(temp[0])
                    break
                else:
                    temp2 = []
                    try:
                        alcohol = driver.find_element(
                            By.XPATH,
                            "//*[@id='wine-page-lower-section']/div[3]/div/div[4]/table/tbody/tr[6]/td/span",
                        ).text
                        temp2.append(alcohol)
                    except:
                        temp2.append("N/A %")
                    for i in temp2:
                        if i[-1] == "%":
                            list.append(temp2[0])
                        else:
                            list.append("N/A %")
                            break
                break

        def find_chi_name():

            wine_name = driver.find_element(
                By.XPATH, "//*[@id='form1']/div[4]/div[2]/h1/div[2]"
            ).text
            temp_name = remove_clickno(wine_name)
            list.append(temp_name + vintage)

        def find_chi_winery():

            wine_chi_winery = driver.find_element(
                By.XPATH, "//*[@id='hlnkWinery']"
            ).text

            list.append(wine_chi_winery)

        def find_chi_describe():

            wine_chi_describe = driver.find_element(
                By.XPATH, "//*[@id='pnlSummary']/div[2]/div"
            ).text

            value = remove_describe_heading(wine_chi_describe)
            if value[:2] == "关于":
                value = "N/A"
            list.append(value)

        def chi_main():
            find_chi_name()
            find_chi_winery()
            find_chi_describe()
            # add vintage
            list.append(vintage)
            print(list)
            sheet["B" + str(i + 2)] = list[0]
            sheet["D" + str(i + 2)] = list[1]
            sheet["F" + str(i + 2)] = list[2]
            sheet["E" + str(i + 2)] = list[3]

        def main1():
            global sheet
            find_wine_name()
            find_category()
            scroll_down()
            time.sleep(2)
            try:
                find_wine_winery()
            except:
                fail_list.append(name + " find_wine_winery has failed")

            try:
                find_graphes()
            except:
                fail_list.append(name + " find_graphes has failed")
            try:
                find_region()
            except:
                fail_list.append(name + " find_region has failed")
            try:
                find_alcohol()
            except:
                fail_list.append(name + " find_alcohol has failed")

            print(list)
            sheet["A" + str(i + 2)] = list[0]
            sheet["G" + str(i + 2)] = list[1]
            sheet["C" + str(i + 2)] = list[2]
            sheet["N" + str(i + 2)] = list[3]
            sheet["P" + str(i + 2)] = list[6]
            sheet["O" + str(i + 2)] = list[4]

            # handle region list
            sheet["H" + str(i + 2)] = list[5][0][4]
            sheet["I" + str(i + 2)] = list[5][0][5]
            sheet["J" + str(i + 2)] = list[5][0][2]
            sheet["K" + str(i + 2)] = list[5][0][3]
            sheet["L" + str(i + 2)] = list[5][0][0]
            sheet["M" + str(i + 2)] = list[5][0][1]

        def main2():
            global sheet

            find_wine_name2()
            find_category()
            scroll_down2()
            time.sleep(2)
            # test(

            find_wine_winery2()

            find_graphes2()

            try:
                find_region2()

            except:
                print(name + " find_region2 has failed")
                fail_list.append(name + " find_region2 has failed")
            find_alcohol2()
            print(list)
            sheet["A" + str(i + 2)] = list[0]
            sheet["G" + str(i + 2)] = list[1]
            sheet["C" + str(i + 2)] = list[2]
            sheet["N" + str(i + 2)] = list[3]
            sheet["P" + str(i + 2)] = list[6]
            sheet["O" + str(i + 2)] = list[4]

            # handle region list
            sheet["H" + str(i + 2)] = list[5][0][4]
            sheet["I" + str(i + 2)] = list[5][0][5]
            sheet["J" + str(i + 2)] = list[5][0][2]
            sheet["K" + str(i + 2)] = list[5][0][3]
            sheet["L" + str(i + 2)] = list[5][0][0]
            sheet["M" + str(i + 2)] = list[5][0][1]

        if k == 0:
            try:
                main2()
            except:
                main1()
        elif k == 1:
            try:
                chi_main()
            except:
                print("chi_main has failed")
                fail_list.append("chi_main has failed")
        else:
            print("The main function has error")

workbook.save("redwineplat.xlsx")
print("fail list :", fail_list)
print()
print("Finish excel data extracting")
print()
